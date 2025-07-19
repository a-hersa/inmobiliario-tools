from flask import Flask, send_file, request, render_template
from src.calculadora import DatosCalculadora
from src.calculadora import calculadora
from src.scraper import scraper
import tempfile
from openpyxl import load_workbook
from dotenv import load_dotenv
import psycopg2
import os
from unidecode import unidecode
import logging
import requests
import json


app=Flask(__name__)
app.secret_key = 'your_secret_key'

# Configurar Flask en modo debug según la variable de entorno
debug_mode = os.getenv("FLASK_DEBUG", "0") == "1"
app.config["DEBUG"] = debug_mode

# Configurar logging
if debug_mode:
    logging.basicConfig(level=logging.DEBUG)
else:
    logging.basicConfig(level=logging.INFO)

# Registra la función para usarla en la plantilla
app.jinja_env.globals.update(calculadora=calculadora)

# Función para verificar Turnstile
def verify_turnstile(token):
    """
    Verifica el token de Turnstile con la API de Cloudflare
    """
    if not token:
        return False
    
    secret_key = os.getenv("TURNSTILE_SECRET_KEY")
    if not secret_key:
        app.logger.error("TURNSTILE_SECRET_KEY not configured")
        # In development mode, allow access if using test site key
        if os.getenv("TURNSTILE_SITE_KEY") == "0x4AAAAAAA49rLAtoXbLsqea":
            app.logger.info("Using test site key in development mode - allowing access")
            return True
        return False
    
    # Datos para enviar a la API de Cloudflare
    data = {
        'secret': secret_key,
        'response': token,
        'remoteip': request.remote_addr
    }
    
    try:
        # Llamada a la API de verificación de Turnstile
        response = requests.post(
            'https://challenges.cloudflare.com/turnstile/v0/siteverify',
            data=data,
            timeout=10
        )
        
        result = response.json()
        
        if result.get('success', False):
            app.logger.info("Turnstile verification successful")
            return True
        else:
            app.logger.warning(f"Turnstile verification failed: {result.get('error-codes', [])}")
            return False
            
    except requests.RequestException as e:
        app.logger.error(f"Turnstile verification request failed: {e}")
        return False
    except json.JSONDecodeError as e:
        app.logger.error(f"Turnstile verification JSON decode failed: {e}")
        return False

# Configurar la conexión a la base de datos PostgreSQL
def get_db_connection():
    # conn = psycopg2.connect(DATABASE_URL)
    conn = psycopg2.connect(
        host='postgres',
        port='5432',
        dbname='inmobiliario_db',
        user= f'{os.getenv("POSTGRES_USER")}',
        password= f'{os.getenv("POSTGRES_PASSWORD")}'
    )
    return conn

@app.route('/')
def home_page():
    app.logger.info("Ruta principal accedida.")

    # Obtener el número de página actual desde los parámetros de consulta
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Elementos por página
    offset = (page - 1) * per_page

    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener los datos paginados incluyendo la columna 'descripcion'
    cursor.execute("""
        SELECT nombre, fecha_updated, precio, metros, poblacion, url, p_id, estatus 
        FROM propiedades 
        WHERE p_id IN (
            SELECT p_id 
            FROM propiedades 
            ORDER BY fecha_updated DESC 
            LIMIT 100
        )
        ORDER BY fecha_updated DESC 
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    propiedades = cursor.fetchall()

    # Obtener el número total de registros
    cursor.execute("SELECT COUNT(*) FROM propiedades")
    total = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    # Calcular el número total de páginas
    total_pages = (total + per_page - 1) // per_page

    # Pasar las propiedades procesadas a la plantilla
    return render_template('index.html', propiedades=propiedades, page=page, total_pages=total_pages, turnstile_site_key=os.getenv("TURNSTILE_SITE_KEY"))


@app.route('/calculadora', methods=['GET', 'POST'])
def calculadora():
    if request.method == 'POST':
        # Verificar Turnstile antes de procesar
        turnstile_token = request.form.get('cf-turnstile-response')
        if not verify_turnstile(turnstile_token):
            app.logger.warning("Turnstile validation failed for POST request")
            return render_template('error.html', 
                                 error_title="Verificación de seguridad fallida",
                                 error_message="Por favor, completa la verificación de seguridad e inténtalo de nuevo."), 400

        # Caso: Solicitud POST con URL del scraper
        url = request.form['url']
        app.logger.info(f"URL recibida: {url}")

        scrape = scraper(url)
        datos = DatosCalculadora(scrape[0], scrape[1], scrape[2], scrape[3]) # nombre, precio, metros, poblacion

        return render_template('calculadora.html', url=url, datos=datos)
    
    elif request.method == 'GET':
        
        # Caso: Solicitud GET con p_id de la base de datos
        p_id = request.args.get('p_id', type=int)
        if p_id is not None:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT nombre, precio, metros, poblacion, url FROM propiedades WHERE p_id = %s", (p_id,))
            propiedad = cursor.fetchone()
            
            conn.close()
            
            if propiedad:
                datos = DatosCalculadora(propiedad[0], propiedad[1], propiedad[2], propiedad[3]) # nombre, precio, metros, poblacion
                return render_template('calculadora.html', url=propiedad[4], datos=datos)
            else:
                return "Propiedad no encontrada", 404
        else:
            # No p_id provided - show blank calculator for manual input
            return render_template('calculadora.html', url="", datos=None)


@app.route('/descargar', methods=['POST'])
def descargar():
    wb = load_workbook('src/plantilla.xlsx')
    ws = wb.active
    ws['D5'] = request.form['nombre']
    ws['D6'] = request.form['url']
    ws['D10'] = int(request.form['p_compra'])
    ws['D11'] = int(request.form['itp'])
    ws['D12'] = int(request.form['reforma'])
    ws['D13'] = int(request.form['notaria'])
    ws['D14'] = int(request.form['registro'])
    ws['D15'] = int(request.form['agencia'])
    ws['D16'] = int(request.form['tasacion'])
    ws['E20'] = int(request.form['ibi'])
    ws['E21'] = int(request.form['basuras'])
    ws['E22'] = int(request.form['comunidad'])
    ws['E23'] = int(request.form['seguros'])
    ws['E24'] = 0
    ws['H5'] = int(request.form['alquiler'])
    ws['H8'] = float(float(request.form['financiado'])/100)
    ws['H11'] = int(request.form['plazo'])
    ws['H12'] = float(float(request.form['intereses_anuales'])/100)
    temp_excel_file = tempfile.NamedTemporaryFile(prefix='calculadora-', suffix='.xlsx', delete=False)
    wb.save(temp_excel_file.name)
    # path = "src/plantilla.xlsx"
    return send_file(temp_excel_file.name, as_attachment=True)

if __name__=='__main__':
    load_dotenv()
    app.run(host='0.0.0.0', port=5000)